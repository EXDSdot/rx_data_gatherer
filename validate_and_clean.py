from __future__ import annotations

import json
import logging
import difflib
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

# --- Helpers ---

def normalize_cik(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if not s or not s.isdigit():
        return None
    return s.zfill(10)

def clean_company_name(name: str) -> str:
    if not name: return ""
    n = str(name).lower()
    # Remove parens content e.g. (2001)
    n = re.sub(r'\s*\(.*?\)', '', n)
    # Remove punctuation
    n = re.sub(r'[^a-z0-9\s]', '', n)
    
    suffixes = [
        'inc', 'incorporated', 'corp', 'corporation', 'llc', 'ltd', 'limited', 
        'plc', 'co', 'company', 'group', 'holdings', 'services', 'systems', 
        'lp', 'partners', 'trust', 'fund'
    ]
    words = n.split()
    cleaned_words = [w for w in words if w not in suffixes]
    return " ".join(cleaned_words)

def get_similarity_score(name_excel: str, name_sec: str) -> float:
    c1 = clean_company_name(name_excel)
    c2 = clean_company_name(name_sec)
    
    if not c1 or not c2: return 0.0
        
    ratio = difflib.SequenceMatcher(None, c1, c2).ratio()
    
    set1 = set(c1.split())
    set2 = set(c2.split())
    intersection = set1.intersection(set2)
    
    is_subset = (len(intersection) == len(set1)) or (len(intersection) == len(set2))
    
    if is_subset and len(intersection) > 0:
        return max(ratio, 0.95)
        
    return ratio

# --- Main ---

def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
    log = logging.getLogger("cleaner")
    
    input_xlsx = "input.xlsx"
    json_dir = Path("json_data")
    
    out_clean = "cleaned_input.xlsx"
    out_review = "_companies_under_consideration.xlsx"
    out_txt = "name_mismatches.txt"  # <--- NEW FILE
    
    # Thresholds
    HIGH_MATCH = 0.80
    REVIEW_THRESHOLD = 0.30  
    TXT_REPORT_THRESHOLD = 0.90   # <--- Report everything below 90% to TXT

    # 1. Check Files
    if not json_dir.exists():
        log.error("JSON folder missing.")
        return
    
    if not Path(input_xlsx).exists():
        log.error("Input Excel missing.")
        return

    # 2. Load Excel
    log.info(f"Reading {input_xlsx}...")
    wb = load_workbook(input_xlsx, read_only=True, data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return

    # 3. Analyze Header
    original_header = list(rows[0])
    header_lower = [str(h).strip().lower() if h else "" for h in original_header]
    
    try:
        cik_idx = -1
        for c in ["cikbefore", "cik", "company_cik"]:
            if c in header_lower: cik_idx = header_lower.index(c); break
        
        name_idx = -1
        for c in ["namecorp", "company_name", "company", "entity_name"]:
            if c in header_lower: name_idx = header_lower.index(c); break

        if cik_idx == -1:
            raise ValueError("Could not find CIK column.")
            
    except ValueError as e:
        log.error(e)
        return

    # 4. Processing
    total_input = len(rows) - 1
    
    valid_rows = []    
    review_rows = []   
    txt_lines = []     # <--- Lines for the text file
    
    seen_ciks = set()
    
    stats = {
        "high_match": 0,    
        "review_match": 0,  
        "low_match": 0,     
        "missing_file": 0,
        "valid_kept": 0
    }

    log.info(f"Processing {total_input} input rows...")

    new_header = original_header + ["sec_entity_name", "name_match_score"]

    for r in rows[1:]:
        if not r: continue
        
        row_values = list(r)
        
        raw_cik = row_values[cik_idx]
        raw_name = row_values[name_idx] if name_idx != -1 else ""

        cik10 = normalize_cik(raw_cik)
        if not cik10 or cik10 == "0000000000": continue
        
        # Check Existence
        json_path = json_dir / f"CIK{cik10}.json"
        if not json_path.exists():
            stats["missing_file"] += 1
            continue

        # Get SEC Name & Match
        sec_name = "N/A"
        score = 0.0
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                sec_name = data.get('entityName', "Unknown")
            
            if raw_name:
                score = get_similarity_score(str(raw_name), sec_name)
        except Exception:
            pass 
        
        # Append info
        row_values.append(sec_name)
        row_values.append(f"{score:.2f}")

        # --- TXT REPORT LOGIC ---
        if score < TXT_REPORT_THRESHOLD:
            # "our name space input name" -> ExcelName SECName
            # Clean newlines just in case to prevent file corruption
            safe_raw = str(raw_name).replace("\n", "").replace("\r", "")
            safe_sec = str(sec_name).replace("\n", "").replace("\r", "")
            txt_lines.append(f"{safe_raw} || {safe_sec}")
        # ------------------------

        # Categorize
        if score >= HIGH_MATCH:
            stats["high_match"] += 1
        elif score >= REVIEW_THRESHOLD:
            stats["review_match"] += 1
            review_rows.append(row_values)
        else:
            stats["low_match"] += 1
        
        valid_rows.append(row_values)
        seen_ciks.add(cik10)
        stats["valid_kept"] += 1

    # 5. Reporting
    log.info("-" * 40)
    log.info(f"Total Input Rows:       {total_input}")
    log.info(f"Dropped (Missing File): {stats['missing_file']}")
    log.info("-" * 40)
    log.info(f"VALID ROWS KEPT:        {stats['valid_kept']}")
    log.info(f"  > High Match (>=80%): {stats['high_match']}")
    log.info(f"  > Review ({int(REVIEW_THRESHOLD*100)}%-80%):   {stats['review_match']}")
    log.info(f"  > Low Match (<{int(REVIEW_THRESHOLD*100)}%):    {stats['low_match']}")
    log.info("-" * 40)

    # 6. Write Outputs
    if stats["valid_kept"] > 0:
        # A) Cleaned Excel
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Cleaned Data"
        ws1.append(new_header)
        for row in valid_rows:
            ws1.append(row)
        wb1.save(out_clean)
        log.info(f"Saved CLEANED DATA to: {out_clean}")

        # B) Review Excel
        if review_rows:
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.title = "Review Needed"
            ws2.append(new_header)
            for row in review_rows:
                ws2.append(row)
            wb2.save(out_review)
            log.info(f"Saved REVIEW CANDIDATES to: {out_review}")
            
        # C) Text File (<90%)
        if txt_lines:
            with open(out_txt, "w", encoding="utf-8") as f:
                for line in txt_lines:
                    f.write(line + "\n")
            log.info(f"Saved {len(txt_lines)} mismatches (<90%) to: {out_txt}")
    else:
        log.error("No valid rows found.")

if __name__ == "__main__":
    main()