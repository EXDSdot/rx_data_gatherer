# import pandas as pd
# import json
# import os
# from dateutil import parser
# from datetime import timedelta
# import numpy as np

# # ================= CONFIGURATION =================
# INPUT_FILE = 'cleaned_input.xlsx'
# OUTPUT_FILE = 'lopucki_xbrl_smart_match.xlsx'
# JSON_FOLDER = './json_data'

# # LoPucki Columns (Input)
# COL_CIK = 'cikbefore'
# COL_DATE = 'date10kbefore'

# # Mappings: (LoPucki Col, XBRL Tag, Is_Duration)
# FIELD_MAP = [
#     {'lopucki': 'assetsbefore',    'xbrl_tag': 'Assets',        'is_duration': False},
#     {'lopucki': 'liabbefore',      'xbrl_tag': 'Liabilities',   'is_duration': False},
#     {'lopucki': 'netincomebefore', 'xbrl_tag': 'NetIncomeLoss', 'is_duration': True}
# ]

# # Settings
# DATE_TOLERANCE_DAYS = 7  # Look for dates within +/- 7 days
# # =================================================

# def normalize_cik(cik_value):
#     try:
#         if pd.isna(cik_value) or cik_value == '':
#             return None
#         return str(int(cik_value)).zfill(10)
#     except:
#         return None

# def parse_date_safe(date_obj):
#     """Returns a datetime object (not string) for comparison."""
#     try:
#         if pd.isna(date_obj) or date_obj == '':
#             return None
#         if isinstance(date_obj, str):
#             return parser.parse(date_obj)
#         return date_obj # Already a timestamp
#     except:
#         return None

# def get_xbrl_value_smart(cik, target_date, tag_name, is_duration=False):
#     """
#     Finds XBRL value with:
#     1. Fuzzy date matching (+/- 7 days).
#     2. Prioritizing 10-K forms.
#     3. Duration check (approx 1 year) for Net Income.
#     """
#     if not cik or not target_date: 
#         return None, "Skipped (Bad Data)"
        
#     filename = f"CIK{cik}.json"
#     filepath = os.path.join(JSON_FOLDER, filename)

#     if not os.path.exists(filepath):
#         return None, "File Not Found"

#     try:
#         with open(filepath, 'r') as f:
#             data = json.load(f)
        
#         # Access US-GAAP facts
#         facts = data.get('facts', {}).get('us-gaap', {})
        
#         # Get the specific tag node
#         tag_node = facts.get(tag_name, {})
#         units = tag_node.get('units', {}).get('USD', [])

#         if not units:
#             return None, f"Tag '{tag_name}' Not Found"

#         # --- SMART MATCHING LOGIC ---
#         candidates = []

#         for entry in units:
#             if 'end' not in entry: continue
            
#             try:
#                 entry_end_date = parser.parse(entry['end'])
                
#                 # Calculate difference in days (Target Date vs XBRL End Date)
#                 diff = abs((entry_end_date - target_date).days)
                
#                 # Check if within tolerance window
#                 if diff <= DATE_TOLERANCE_DAYS:
                    
#                     # --- Duration Check for Net Income ---
#                     if is_duration:
#                         if 'start' not in entry: continue
#                         entry_start_date = parser.parse(entry['start'])
#                         duration_days = (entry_end_date - entry_start_date).days
#                         # We only want Annual numbers (approx 365 days)
#                         # Rejecting if duration is too short (e.g. 90 days for Q4)
#                         if not (350 <= duration_days <= 375):
#                             continue 
                    
#                     form_type = entry.get('form', '').upper()
#                     val = entry.get('val')
                    
#                     # Store candidate details
#                     candidates.append({
#                         'val': val,
#                         'diff': diff,
#                         'date': entry['end'],
#                         'form': form_type,
#                         'is_10k': '10-K' in form_type
#                     })
#             except:
#                 continue # Skip bad dates in JSON

#         if not candidates:
#             return None, f"No match within {DATE_TOLERANCE_DAYS} days"

#         # --- SELECTION STRATEGY ---
#         # Sort by: 
#         # 1. Is it a 10-K? (True comes first)
#         # 2. Date difference (Smallest diff is best)
        
#         candidates.sort(key=lambda x: (not x['is_10k'], x['diff']))
        
#         best = candidates[0]
        
#         #match_note = f"Found {best['date']} (Diff: {best['diff']}d, Form: {best['form']})"
#         match_note="found"
#         return best['val'], match_note

#     except Exception as e:
#         return None, f"Error: {str(e)}"

# def main():
#     print("Loading input Excel file...")
#     try:
#         df = pd.read_excel(INPUT_FILE)
#     except FileNotFoundError:
#         print(f"Error: Could not find {INPUT_FILE}")
#         return

#     # Normalize columns
#     df.columns = df.columns.str.strip().str.lower()
    
#     print(f"Processing {len(df)} rows with Smart Matching...")
    
#     # Store results in lists of dicts to easily convert to DataFrame later
#     results_data = []

#     for index, row in df.iterrows():
#         cik_clean = normalize_cik(row.get(COL_CIK))
#         date_obj = parse_date_safe(row.get(COL_DATE))
        
#         row_result = row.to_dict() # Start with existing data
        
#         # We track if the primary asset lookup failed to determine sheet placement
#         primary_lookup_failed = False
        
#         for field in FIELD_MAP:
#             l_col = field['lopucki']
#             x_tag = field['xbrl_tag']
#             is_dur = field['is_duration']
            
#             # Fetch Data
#             val, info = get_xbrl_value_smart(cik_clean, date_obj, x_tag, is_dur)
            
#             # Save Raw XBRL Value and Note
#             row_result[f'{x_tag}_XBRL'] = val
#             row_result[f'{x_tag}_Note'] = info
            
#             # Calculate Diff % (if both values exist)
#             lopucki_val = row.get(l_col)
            
#             if val is not None and pd.notna(lopucki_val) and lopucki_val != 0:
#                 # Lopucki is usually in Millions, XBRL in Units.
#                 # Adjust XBRL to Millions:
#                 xbrl_mil = val / 1_000_000
#                 diff_pct = (xbrl_mil - lopucki_val) / lopucki_val
#                 row_result[f'{x_tag}_Diff_%'] = diff_pct
#             else:
#                 row_result[f'{x_tag}_Diff_%'] = None

#             # Mark as failed if Assets (primary) was not found
#             if x_tag == 'Assets' and val is None:
#                 primary_lookup_failed = True

#         row_result['Match_Failed'] = primary_lookup_failed
#         results_data.append(row_result)

#         if index % 100 == 0:
#             print(f"Processed {index} rows...")

#     # Create final DataFrame
#     final_df = pd.DataFrame(results_data)

#     # --- SPLIT SHEETS ---
#     # Sheet 1: Matched (Clean) - Where Match_Failed is False
#     df_matched = final_df[final_df['Match_Failed'] == False].drop(columns=['Match_Failed'])
    
#     # Sheet 2: Unmatched / Errors - Where Match_Failed is True
#     df_unmatched = final_df[final_df['Match_Failed'] == True].drop(columns=['Match_Failed'])

#     print(f"Saving to {OUTPUT_FILE}...")
#     print(f"Matched Rows: {len(df_matched)}")
#     print(f"Unmatched Rows: {len(df_unmatched)}")

#     with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
#         df_matched.to_excel(writer, sheet_name='Matched_Data', index=False)
#         df_unmatched.to_excel(writer, sheet_name='Unmatched_Errors', index=False)
        
#     print("Done!")

# if __name__ == "__main__":
#     main()
import pandas as pd
import json
import os
from dateutil import parser
import numpy as np
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# ================= CONFIGURATION =================
INPUT_FILE = 'cleaned_input.xlsx'
OUTPUT_FILE = 'lopucki_xbrl_smart_match.xlsx'
JSON_FOLDER = './json_data'

# LoPucki Columns (Input)
COL_CIK = 'cikbefore'
COL_DATE = 'date10kbefore'

DATE_TOLERANCE_DAYS = 7  # Look for dates within +/- 7 days

# XBRL Fields to pull (base items -> ratios computed later)
# kind: "instant" = balance sheet point-in-time, "duration" = annual flow (needs start/end ~ 1y)
FIELD_MAP = [
    # Balance sheet (instant)
    {"xbrl_tag": "Assets", "kind": "instant", "lopucki": "assetsbefore"},
    {"xbrl_tag": "Liabilities", "kind": "instant", "lopucki": "liabbefore"},
    {"xbrl_tag": "AssetsCurrent", "kind": "instant"},
    {"xbrl_tag": "LiabilitiesCurrent", "kind": "instant"},
    {"xbrl_tag": "CashAndCashEquivalentsAtCarryingValue", "kind": "instant"},
    {"xbrl_tag": "DebtCurrent", "kind": "instant"},
    {"xbrl_tag": "LongTermDebt", "kind": "instant"},
    {"xbrl_tag": "StockholdersEquity", "kind": "instant"},
    {"xbrl_tag": "InventoryNet", "kind": "instant"},
    {"xbrl_tag": "AccountsReceivableNetCurrent", "kind": "instant"},
    {"xbrl_tag": "PropertyPlantAndEquipmentNet", "kind": "instant"},
    {"xbrl_tag": "Goodwill", "kind": "instant"},
    {"xbrl_tag": "IntangibleAssetsNetExcludingGoodwill", "kind": "instant"},

    # Income statement / cash flow (duration - annual)
    {"xbrl_tag": "Revenues", "kind": "duration"},
    {"xbrl_tag": "CostOfRevenue", "kind": "duration"},
    {"xbrl_tag": "SellingGeneralAndAdministrativeExpense", "kind": "duration"},
    {"xbrl_tag": "OperatingIncomeLoss", "kind": "duration"},
    {"xbrl_tag": "InterestExpense", "kind": "duration"},
    {"xbrl_tag": "NetIncomeLoss", "kind": "duration", "lopucki": "netincomebefore"},
    {"xbrl_tag": "NetCashProvidedByUsedInOperatingActivities", "kind": "duration"},
    {"xbrl_tag": "PaymentsToAcquirePropertyPlantAndEquipment", "kind": "duration"},
]
# =================================================


def normalize_cik(cik_value):
    try:
        if pd.isna(cik_value) or cik_value == '':
            return None
        return str(int(cik_value)).zfill(10)
    except:
        return None


def parse_date_safe(date_obj):
    try:
        if pd.isna(date_obj) or date_obj == '':
            return None
        if isinstance(date_obj, str):
            return parser.parse(date_obj)
        return date_obj
    except:
        return None


def _get_units_list(tag_node, preferred_unit="USD"):
    """
    SEC companyfacts JSON often has units like USD, shares, pure, USD/shares, etc.
    We prefer USD; fallback to the first available unit list.
    """
    units = tag_node.get("units", {})
    if preferred_unit in units and units[preferred_unit]:
        return units[preferred_unit], preferred_unit

    # fallback: pick first non-empty units list
    for unit_key, unit_list in units.items():
        if unit_list:
            return unit_list, unit_key
    return [], None


def get_xbrl_value_smart(cik, target_date, tag_name, kind="instant"):
    """
    Smart match:
      - end date within +/- tolerance
      - prefer 10-K
      - for duration items: require ~ annual (350-375 days)
    Returns: value (float/int) or None
    """
    if not cik or not target_date:
        return None

    filepath = os.path.join(JSON_FOLDER, f"CIK{cik}.json")
    if not os.path.exists(filepath):
        return None

    try:
        with open(filepath, "r") as f:
            data = json.load(f)

        facts = data.get("facts", {}).get("us-gaap", {})
        tag_node = facts.get(tag_name, {})
        if not tag_node:
            return None

        unit_list, unit_key = _get_units_list(tag_node, preferred_unit="USD")
        if not unit_list:
            return None

        candidates = []
        for entry in unit_list:
            if "end" not in entry:
                continue

            try:
                entry_end = parser.parse(entry["end"])
                diff_days = abs((entry_end - target_date).days)
                if diff_days > DATE_TOLERANCE_DAYS:
                    continue

                if kind == "duration":
                    if "start" not in entry:
                        continue
                    entry_start = parser.parse(entry["start"])
                    dur = (entry_end - entry_start).days
                    if not (350 <= dur <= 375):
                        continue

                form_type = (entry.get("form") or "").upper()
                val = entry.get("val", None)
                if val is None:
                    continue

                candidates.append({
                    "val": val,
                    "diff": diff_days,
                    "is_10k": ("10-K" in form_type),
                })
            except:
                continue

        if not candidates:
            return None

        candidates.sort(key=lambda x: (not x["is_10k"], x["diff"]))
        return candidates[0]["val"]

    except:
        return None


def safe_div(a, b):
    if a is None or b is None:
        return None
    try:
        if pd.isna(a) or pd.isna(b) or b == 0:
            return None
        return a / b
    except:
        return None


def as_num(x):
    if x is None:
        return None
    try:
        if pd.isna(x):
            return None
        return float(x)
    except:
        return None


def apply_green_red_fills(excel_path, sheet_names, target_col_suffixes):
    """
    Colors:
      - green if cell has a value
      - red if blank/NaN
    Only applies to columns whose header ends with one of target_col_suffixes.
    """
    wb = load_workbook(excel_path)
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for ws_name in sheet_names:
        if ws_name not in wb.sheetnames:
            continue
        ws = wb[ws_name]
        if ws.max_row < 2 or ws.max_column < 1:
            continue

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        target_cols = []
        for idx, h in enumerate(headers, start=1):
            if isinstance(h, str) and any(h.endswith(suf) for suf in target_col_suffixes):
                target_cols.append(idx)

        for r in range(2, ws.max_row + 1):
            for c in target_cols:
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is None or (isinstance(v, float) and np.isnan(v)) or v == "":
                    cell.fill = red
                else:
                    cell.fill = green

    wb.save(excel_path)


def main():
    df = pd.read_excel(INPUT_FILE)
    df.columns = df.columns.str.strip().str.lower()

    results = []

    for _, row in df.iterrows():
        cik = normalize_cik(row.get(COL_CIK))
        target_date = parse_date_safe(row.get(COL_DATE))

        out = row.to_dict()

        # Pull raw XBRL values
        for field in FIELD_MAP:
            tag = field["xbrl_tag"]
            kind = field["kind"]
            val = get_xbrl_value_smart(cik, target_date, tag, kind=kind)

            out[f"{tag}_XBRL"] = val

            # Optional validation vs LoPucki if provided
            lop_col = field.get("lopucki")
            if lop_col:
                lop_val = row.get(lop_col)
                val_num = as_num(val)
                if val_num is not None and pd.notna(lop_val) and lop_val != 0:
                    xbrl_mil = val_num / 1_000_000
                    out[f"{tag}_Diff_%"] = (xbrl_mil - lop_val) / lop_val
                else:
                    out[f"{tag}_Diff_%"] = None

        # ---- Compute ratios (use the XBRL columns) ----
        A = as_num(out.get("Assets_XBRL"))
        L = as_num(out.get("Liabilities_XBRL"))
        AC = as_num(out.get("AssetsCurrent_XBRL"))
        LC = as_num(out.get("LiabilitiesCurrent_XBRL"))
        CASH = as_num(out.get("CashAndCashEquivalentsAtCarryingValue_XBRL"))
        Dcur = as_num(out.get("DebtCurrent_XBRL"))
        Dlt = as_num(out.get("LongTermDebt_XBRL"))
        REV = as_num(out.get("Revenues_XBRL"))
        COGS = as_num(out.get("CostOfRevenue_XBRL"))
        SGA = as_num(out.get("SellingGeneralAndAdministrativeExpense_XBRL"))
        OI = as_num(out.get("OperatingIncomeLoss_XBRL"))
        INT = as_num(out.get("InterestExpense_XBRL"))
        NI = as_num(out.get("NetIncomeLoss_XBRL"))
        CFO = as_num(out.get("NetCashProvidedByUsedInOperatingActivities_XBRL"))
        CAPEX = as_num(out.get("PaymentsToAcquirePropertyPlantAndEquipment_XBRL"))
        PPE = as_num(out.get("PropertyPlantAndEquipmentNet_XBRL"))
        GW = as_num(out.get("Goodwill_XBRL"))
        INTANG = as_num(out.get("IntangibleAssetsNetExcludingGoodwill_XBRL"))
        INV = as_num(out.get("InventoryNet_XBRL"))
        AR = as_num(out.get("AccountsReceivableNetCurrent_XBRL"))
        EQ = as_num(out.get("StockholdersEquity_XBRL"))

        debt = None if (Dcur is None and Dlt is None) else (0 if Dcur is None else Dcur) + (0 if Dlt is None else Dlt)
        intangibles = None if (GW is None and INTANG is None) else (0 if GW is None else GW) + (0 if INTANG is None else INTANG)

        # EBITDA proxy (simple & common): REV - COGS - SGA
        EBITDA = None
        if REV is not None and COGS is not None and SGA is not None:
            EBITDA = REV - COGS - SGA

        out["Leverage_Liab_to_Assets_R"] = safe_div(L, A)
        out["Debt_to_Assets_R"] = safe_div(debt, A)
        out["Cash_to_Assets_R"] = safe_div(CASH, A)
        out["Current_Ratio_R"] = safe_div(AC, LC)
        out["WorkingCap_to_Assets_R"] = safe_div((AC - LC) if (AC is not None and LC is not None) else None, A)

        out["ROA_NI_to_Assets_R"] = safe_div(NI, A)
        out["OperatingMargin_OI_to_Sales_R"] = safe_div(OI, REV)
        out["GrossMargin_R"] = safe_div((REV - COGS) if (REV is not None and COGS is not None) else None, REV)
        out["AssetTurnover_Sales_to_Assets_R"] = safe_div(REV, A)

        out["EBITDA_Margin_R"] = safe_div(EBITDA, REV)
        out["InterestCoverage_OI_to_IntExp_R"] = safe_div(OI, INT)

        out["CFO_to_Assets_R"] = safe_div(CFO, A)
        out["FCF_to_Assets_R"] = safe_div((CFO - CAPEX) if (CFO is not None and CAPEX is not None) else None, A)
        out["Accruals_NI_minus_CFO_to_Assets_R"] = safe_div((NI - CFO) if (NI is not None and CFO is not None) else None, A)

        out["PPE_to_Assets_R"] = safe_div(PPE, A)
        out["Intangibles_to_Assets_R"] = safe_div(intangibles, A)
        out["Inventory_to_Assets_R"] = safe_div(INV, A)
        out["AR_to_Sales_R"] = safe_div(AR, REV)
        out["BookLeverage_Debt_to_Equity_R"] = safe_div(debt, EQ)

        results.append(out)

    final_df = pd.DataFrame(results)

    # Split sheets the same way you were doing, but simpler:
    # treat "Assets missing" as a failed match
    final_df["Match_Failed"] = final_df["Assets_XBRL"].isna()
    df_matched = final_df[~final_df["Match_Failed"]].drop(columns=["Match_Failed"])
    df_unmatched = final_df[final_df["Match_Failed"]].drop(columns=["Match_Failed"])

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df_matched.to_excel(writer, sheet_name="Matched_Data", index=False)
        df_unmatched.to_excel(writer, sheet_name="Unmatched_Errors", index=False)

    # Color XBRL + Ratio columns (green present, red missing)
    apply_green_red_fills(
        OUTPUT_FILE,
        sheet_names=["Matched_Data", "Unmatched_Errors"],
        target_col_suffixes=("_XBRL", "_R")  # XBRL fields + ratios
    )

    print("Done:", OUTPUT_FILE)


if __name__ == "__main__":
    main()