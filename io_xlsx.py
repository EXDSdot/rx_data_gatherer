from __future__ import annotations

import logging
import os
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


log = logging.getLogger("io_xlsx")


def _to_iso_date(val: Any) -> str | None:
    if val is None:
        return None

    if isinstance(val, (datetime, date)):
        return val.date().isoformat() if isinstance(val, datetime) else val.isoformat()

    s = str(val).strip()
    if not s:
        return None

    # already ISO
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]

    # common formats
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass

    return None


def load_cik_event_dates_xlsx(path: str, sheet_name: str | None = None) -> list[tuple[str, str]]:
    """
    Reads .xlsx with columns for CIK and Event Date.
    Supports standard names (CIK, start_date) and LoPucki BRD names (CikBefore, DateFiled).
    """
    if not os.path.exists(path):
        log.error("Excel file not found: %s (cwd=%s)", path, os.getcwd())
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Read header
    row_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        log.error("Excel file %s is empty", path)
        return []

    header = [str(x).strip().lower() if x is not None else "" for x in header_row]

    def find_col(names: set[str]) -> int | None:
        for i, h in enumerate(header):
            if h in names:
                return i
        return None

    # Expanded candidates for BRD support
    # Note: 'or 0' / 'or 1' fallbacks are kept for legacy/headerless support, 
    # but specific headers will take precedence.
    cik_col = find_col({"cikbefore", "cik_before"})
    if cik_col is None:
        cik_col = 0
        
    start_col = find_col({
        "datefiled", "date_filed"
    })
    if start_col is None:
        start_col = 1

    out: list[tuple[str, str]] = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
            
        # Ensure row is long enough
        max_idx = max(cik_col, start_col)
        if len(row) <= max_idx:
            continue

        cik_raw = row[cik_col]
        start_raw = row[start_col]

        if cik_raw is None:
            skipped += 1
            continue

        cik_str = str(cik_raw).strip()
        if not cik_str:
            skipped += 1
            continue

        # Check for non-date placeholders often found in raw data
        if start_raw is None or str(start_raw).strip() == "":
            skipped += 1
            continue

        event_iso = _to_iso_date(start_raw)
        if not event_iso:
            skipped += 1
            continue

        out.append((cik_str, event_iso))

    log.info("Loaded %d rows (skipped %d) from %s", len(out), skipped, path)
    return out


def write_rx_snapshot_xlsx(results: list[dict[str, Any]], path: str) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "rx_snapshot"

    headers = [
        "cik", "entityName", "event_date", "report_end",
        "age_days", "report_form", "report_fp", "report_filed",
        "coverage", "has_companyfacts",

        # base values + tag used
        "cash_val", "cash_tag",
        "liab_val", "liab_tag",
        "assets_val", "assets_tag",
        "assets_cur_val", "assets_cur_tag",
        "liab_cur_val", "liab_cur_tag",
        "ar_val", "ar_tag",
        "inv_val", "inv_tag",
        "debt_val", "debt_tag",
        "oi_val", "oi_tag",
        "int_val", "int_tag",
        "ocf_val", "ocf_tag",

        # 6 metrics
        "cash_to_liab",
        "current_ratio",
        "quick_ratio",
        "debt_to_assets",
        "interest_coverage",
        "ocf_to_debt",

        "error",
    ]
    ws.append(headers)

    # header style
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.alignment = Alignment(horizontal="center", vertical="center")

    def num(x: Any) -> float | None:
        if x is None or x == "":
            return None
        if isinstance(x, (int, float)):
            return float(x)
        try:
            return float(x)
        except Exception:
            return None

    for r in results:
        meta = r.get("report_meta") or {}
        ws.append([
            r.get("cik", ""),
            r.get("entityName", ""),
            r.get("event_date", ""),
            r.get("report_end", ""),

            meta.get("age_days"),
            meta.get("form"),
            meta.get("fp"),
            meta.get("filed"),
            meta.get("coverage"),
            int(r.get("has_companyfacts", 0) or 0),

            num(r.get("cash_val")), r.get("cash_tag"),
            num(r.get("liab_val")), r.get("liab_tag"),
            num(r.get("assets_val")), r.get("assets_tag"),
            num(r.get("assets_cur_val")), r.get("assets_cur_tag"),
            num(r.get("liab_cur_val")), r.get("liab_cur_tag"),
            num(r.get("ar_val")), r.get("ar_tag"),
            num(r.get("inv_val")), r.get("inv_tag"),
            num(r.get("debt_val")), r.get("debt_tag"),
            num(r.get("oi_val")), r.get("oi_tag"),
            num(r.get("int_val")), r.get("int_tag"),
            num(r.get("ocf_val")), r.get("ocf_tag"),

            num(r.get("cash_to_liab")),
            num(r.get("current_ratio")),
            num(r.get("quick_ratio")),
            num(r.get("debt_to_assets")),
            num(r.get("interest_coverage")),
            num(r.get("ocf_to_debt")),

            r.get("error", ""),
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # widths
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        if h in {"entityName", "error"}:
            ws.column_dimensions[col].width = 40
        elif h in {"cik"}:
            ws.column_dimensions[col].width = 12
        elif h.endswith("_tag") or h.startswith("report_"):
            ws.column_dimensions[col].width = 22
        else:
            ws.column_dimensions[col].width = 16

    # number formats
    val_cols = [h for h in headers if h.endswith("_val")]
    ratio_cols = ["cash_to_liab", "current_ratio", "quick_ratio", "debt_to_assets", "interest_coverage", "ocf_to_debt"]

    for row in range(2, ws.max_row + 1):
        for col_name in val_cols:
            cidx = headers.index(col_name) + 1
            cell = ws.cell(row=row, column=cidx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
        for col_name in ratio_cols:
            cidx = headers.index(col_name) + 1
            cell = ws.cell(row=row, column=cidx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"

    wb.save(p)
    log.info("Wrote %s (%d rows)", str(p), len(results))


def load_court_cases_xlsx(path: str, sheet_name: str | None = None) -> list[dict[str, str]]:
    """
    Expected headers (recommended):
      cik | court | docket_number | filed_date

    Returns list of dict rows:
      {"cik": "...", "court": "...", "docket_number": "...", "filed_date": "..."}
    """
    log = logging.getLogger("court_xlsx_loader")

    if not os.path.exists(path):
        log.error("Excel file not found: %s (cwd=%s)", path, os.getcwd())
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    log.info("Using sheet: %s", ws.title)

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []

    header = [str(x).strip().lower() if x is not None else "" for x in header_row]

    def find_col(cands: set[str]) -> int | None:
        for i, h in enumerate(header):
            if h in cands:
                return i
        return None

    cik_col = find_col({"cik", "cik10", "cikbefore", "cik_before"})
    court_col = find_col({"court", "court_code"})
    docket_col = find_col({"docket_number", "docket", "case_number", "case_no", "case"})
    filed_col = find_col({"filed_date", "filed", "petition_date", "date_filed", "datefiled"})

    if cik_col is None or court_col is None or docket_col is None or filed_col is None:
        log.error(
            "Missing required headers. Need: cik, court, docket_number, filed_date. Got: %s",
            header,
        )
        return []

    out: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(cik_col, court_col, docket_col, filed_col):
            continue

        cik = str(row[cik_col]).strip() if row[cik_col] is not None else ""
        court = str(row[court_col]).strip() if row[court_col] is not None else ""
        docket_number = str(row[docket_col]).strip() if row[docket_col] is not None else ""
        filed_date = _to_iso_date(row[filed_col])

        if not court or not docket_number or not filed_date:
            continue

        out.append(
            {
                "cik": cik,
                "court": court,
                "docket_number": docket_number,
                "filed_date": filed_date,
            }
        )

    return out


def write_court_metrics_xlsx(results: list[dict[str, Any]], path: str = "court_metrics.xlsx") -> None:
    log = logging.getLogger("write_court_metrics_xlsx")
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    # dynamic day columns (collect all keys like docket_count_90d, motion_count_90d)
    day_keys: list[str] = []
    if results:
        keys = set().union(*(r.keys() for r in results if isinstance(r, dict)))
        day_keys = sorted([k for k in keys if k.startswith("docket_count_") or k.startswith("motion_count_")])

    headers = [
        "cik",
        "court",
        "docket_number",
        "filed_date",
        "found",
        "docket_id",
        "total_entries_loaded",
        *day_keys,
        "error",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "court_metrics"
    ws.append(headers)

    # header styling
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        if not isinstance(r, dict):
            continue
        ws.append([r.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # column widths
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        if h in {"error"}:
            ws.column_dimensions[col].width = 45
        elif h in {"docket_number"}:
            ws.column_dimensions[col].width = 22
        else:
            ws.column_dimensions[col].width = 16

    wb.save(p)
    log.info("Wrote %s (%d rows)", str(p), len(results))