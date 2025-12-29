from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def normalize_cik_str(v: Any) -> str:
    s = str(v).strip()
    if not s:
        return ""
    # Try to make it an integer string, then zfill
    try:
        # remove float .0 if exists
        f = float(s)
        i = int(f)
        return str(i).zfill(10)
    except ValueError:
        return s.upper()

def load_brd_map(path: str) -> tuple[list[str], dict[str, dict[str, Any]]]:
    """
    Returns (list_of_columns, map_cik_to_row_dict)
    """
    log = logging.getLogger("post_merge")
    p = Path(path)
    if not p.exists():
        log.warning("LoPucki BRD file not found at %s. Skipping merge.", path)
        return [], {}

    # read_only=True for performance on larger datasets
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    
    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return [], {}

    # Find key column "cikbefore" (case insensitive)
    key_idx = -1
    clean_headers = []
    for i, h in enumerate(headers):
        h_str = str(h).strip() if h is not None else f"col_{i}"
        clean_headers.append(h_str)
        if h_str.lower() == "cikbefore":
            key_idx = i
    
    if key_idx == -1:
        log.warning("Column 'cikbefore' not found in %s. Found: %s", path, clean_headers)
        return [], {}

    # Map CIK -> Row Data
    data_map: dict[str, dict[str, Any]] = {}
    
    count = 0
    for row in rows:
        if not row:
            continue
        
        # Get key
        if key_idx < len(row):
            raw_key = row[key_idx]
            if raw_key is None:
                continue
            cik_norm = normalize_cik_str(raw_key)
            if cik_norm == "0000000000":
                continue
            
            # Store row data mapped by header name
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(clean_headers):
                    row_dict[clean_headers[i]] = val
            
            data_map[cik_norm] = row_dict
            count += 1

    log.info("Loaded %d rows from LoPucki BRD (%s)", count, path)
    return clean_headers, data_map


def merge_lopucki_to_features(
    features_path: str,
    lopucki_path: str,
    output_path: str = ""
) -> str:
    """
    Merges LoPucki columns into the features file. Returns the path to the merged file.
    """
    log = logging.getLogger("post_merge")
    
    brd_headers, brd_map = load_brd_map(lopucki_path)
    if not brd_map:
        return features_path

    # We load the FEATURES file
    fp = Path(features_path)
    if not fp.exists():
        log.error("Features file not found: %s", features_path)
        return features_path
        
    wb_main = load_workbook(fp)
    ws_main = wb_main.active
    
    # Read all rows into memory to append columns safely
    rows = list(ws_main.iter_rows(values_only=False)) 
    
    if not rows:
        return features_path

    header_row = rows[0]
    headers_main = [str(c.value).strip() for c in header_row]
    
    # Find "cik" index in our features file
    cik_idx = -1
    for i, h in enumerate(headers_main):
        if h.lower() == "cik":
            cik_idx = i
            break
            
    if cik_idx == -1:
        log.error("CIK column not found in features file.")
        return features_path

    # Determine columns to add (exclude key, handle name collisions)
    existing_lower = {h.lower() for h in headers_main}
    cols_to_add = []
    
    # Keep track of the mapping from BRD header -> New Column Name
    # (Though we iterate the map directly below)

    for h in brd_headers:
        new_name = h
        if h.lower() in existing_lower or h.lower() == "cikbefore":
            new_name = f"brd_{h}"
        
        cols_to_add.append(new_name)
    
    # Append headers to the first row
    start_col = len(headers_main) + 1
    for i, col_name in enumerate(cols_to_add):
        c = ws_main.cell(row=1, column=start_col + i)
        c.value = col_name
    
    # Iterate data rows (starting from row 2)
    matched = 0
    for r_idx in range(1, len(rows)):
        row_cells = rows[r_idx]
        
        if cik_idx >= len(row_cells):
            continue
            
        cik_val = row_cells[cik_idx].value
        cik_norm = normalize_cik_str(cik_val)
        
        if cik_norm in brd_map:
            row_data = brd_map[cik_norm]
            matched += 1
            
            current_write_col = start_col
            for h in brd_headers:
                val = row_data.get(h)
                c = ws_main.cell(row=r_idx + 1, column=current_write_col)
                c.value = val
                current_write_col += 1

    log.info("Merged LoPucki data into %d rows.", matched)
    
    out = output_path or features_path.replace(".xlsx", "_merged.xlsx")
    wb_main.save(out)
    log.info("Saved merged file to %s", out)
    return out


# ---------------------------------------------------------
# Regression Dataset Generation
# ---------------------------------------------------------
def _clean_str(x: Any) -> str:
    return str(x).strip().lower() if x is not None else ""

def _transform_ceo(val: Any) -> int | None:
    # 1 if "Replaced", 0 if "NoReplace"
    s = _clean_str(val)
    if "noreplace" in s:
        return 0
    if "replaced" in s:
        return 1
    return None

def _transform_chapter(val: Any) -> int | None:
    # 0 if 7, 1 if 11
    s = _clean_str(val)
    if s == "7": return 0
    if s == "11": return 1
    return None

def _transform_yes_no(val: Any) -> int | None:
    # 1 if yes, 2 if no
    s = _clean_str(val)
    if s == "yes": return 1
    if s == "no": return 2
    return None

def _transform_prepackaged(val: Any) -> int | None:
    # 1 if "free fall", 2 if "not applicable", 3 if "prenegotiated"
    s = _clean_str(val)
    if s == "free fall": return 1
    if s == "not applicable": return 2
    if s == "prenegotiated": return 3
    return None

def _transform_voluntary(val: Any) -> int | None:
    # 1 if "voluntary", 2 if "involuntary", 3 if "both"
    s = _clean_str(val)
    if s == "voluntary": return 1
    if s == "involuntary": return 2
    if s == "both": return 3
    return None

def generate_regression_file(merged_path: str, out_path: str) -> None:
    log = logging.getLogger("post_merge")
    if not Path(merged_path).exists():
        log.warning("Merged file not found for regression gen: %s", merged_path)
        return

    wb_in = load_workbook(merged_path, data_only=True)
    ws_in = wb_in.active
    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        return

    headers_in = [str(h).strip() for h in rows[0]]
    
    # Map lowercase column name -> index
    # We must handle potential 'brd_' prefix by checking matches
    col_map = {h.lower(): i for i, h in enumerate(headers_in)}
    
    def find_col_idx(target: str) -> int | None:
        t = target.lower()
        # 1. Exact match
        if t in col_map:
            return col_map[t]
        # 2. 'brd_' match
        if f"brd_{t}" in col_map:
            return col_map[f"brd_{t}"]
        return None

    # --- Definition of Output Columns ---
    # (Source Name, Output Name, Transform Func)
    # Note: incomebebefore -> handled as IncomeBefore if not found
    
    cols_to_copy = [
        "assetsbefore", "daysin", "ebitbefore", "emplbefore",
        "incomebebefore", "intercompanypct", "liabbefore", 
        "netincomebefore", "filingrate"
    ]
    
    # Setup transform map
    # Using list of tuples: (InputName, OutputName, TransformFunc)
    transforms = [
        ("CeoReplaced", "CeoReplaced", _transform_ceo),
        ("chapter", "chapter", _transform_chapter),
        ("claimsagent", "claimsagent", _transform_yes_no),
        ("commcred", "commcred", _transform_yes_no),
        ("prepackaged", "prepackaged", _transform_prepackaged),
        ("voluntary", "voluntary", _transform_voluntary),
    ]

    # Resolve indices
    # We include CIK and EntityName for reference
    ref_cols = ["cik", "entityName"]
    
    final_cols = [] # List of (SourceIdx, OutputName, TransformFunc or None)
    
    # 1. Reference Cols
    for rc in ref_cols:
        idx = find_col_idx(rc)
        if idx is not None:
            final_cols.append((idx, rc, None))
    
    # 2. Direct Copy Cols
    for cc in cols_to_copy:
        idx = find_col_idx(cc)
        if idx is None and cc == "incomebebefore":
            # Typo fallback
            idx = find_col_idx("incomebefore")
        
        final_cols.append((idx, cc, None)) # idx might be None -> write blank

    # 3. Transformed Cols
    for src_name, out_name, func in transforms:
        idx = find_col_idx(src_name)
        final_cols.append((idx, out_name, func))

    # --- Write Output ---
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "regression_data"

    # Write Header
    out_headers = [x[1] for x in final_cols]
    ws_out.append(out_headers)
    
    # Style Header
    for i in range(1, len(out_headers) + 1):
        c = ws_out.cell(row=1, column=i)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
        c.alignment = Alignment(horizontal="center")

    # Write Data
    for r in rows[1:]: # skip header
        out_row = []
        for src_idx, _, func in final_cols:
            val = r[src_idx] if src_idx is not None and src_idx < len(r) else None
            
            if func:
                val = func(val)
            
            out_row.append(val)
        ws_out.append(out_row)

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = f"A1:{get_column_letter(len(out_headers))}1"
    
    wb_out.save(out_path)
    log.info("Saved regression analysis file to %s", out_path)