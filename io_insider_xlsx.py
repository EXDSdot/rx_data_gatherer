from __future__ import annotations

import logging
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _to_iso_date(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()

    s = str(val).strip()
    if not s:
        return None
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]

    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def load_cik_event_dates_xlsx(path: str, sheet_name: str | None = None) -> list[tuple[str, str]]:
    """
    Input: Excel with at least:
      - cik
      - start_date OR event_date

    Returns: [(cik_raw, event_iso), ...]
    """
    log = logging.getLogger("insider_xlsx_loader")

    if not os.path.exists(path):
        log.error("Excel file not found: %s (cwd=%s)", path, os.getcwd())
        return []

    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(x).strip().lower() if x is not None else "" for x in header_row]

    def find_col(names: set[str]) -> int | None:
        for i, h in enumerate(header):
            if h in names:
                return i
        return None

    cik_col = find_col({"cik", "cik10", "company_cik"}) or 0
    event_col = (
        find_col({"event_date", "event", "start_date", "start", "filed_date", "petition_date"}) or 1
    )

    out: list[tuple[str, str]] = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) <= max(cik_col, event_col):
            continue

        cik_raw = row[cik_col]
        event_raw = row[event_col]

        if cik_raw is None or event_raw is None:
            skipped += 1
            continue

        event_iso = _to_iso_date(event_raw)
        if not event_iso:
            skipped += 1
            continue

        out.append((str(cik_raw).strip(), event_iso))

    log.info("Parsed %d rows (skipped %d)", len(out), skipped)
    return out


def write_insider_snapshot_xlsx(results: list[dict[str, Any]], path: str) -> None:
    log = logging.getLogger("write_insider_snapshot_xlsx")
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "insider"

    headers = [
        "cik",
        "entityName",
        "event_date",
        "lookback_days",
        "form4_filings",
        "tx_count",
        "buy_tx",
        "sell_tx",
        "net_shares",
        "net_value_usd",
        "unique_insiders",
        "first_tx_date",
        "last_tx_date",
        "error",
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.alignment = Alignment(horizontal="center", vertical="center")

    def num(x: Any) -> float | None:
        try:
            if x is None or x == "":
                return None
            return float(x)
        except Exception:
            return None

    for r in results:
        ws.append(
            [
                r.get("cik", ""),
                r.get("entityName", "") or "",
                r.get("event_date", "") or "",
                int(r.get("lookback_days", 0) or 0),
                int(r.get("form4_filings", 0) or 0),
                int(r.get("tx_count", 0) or 0),
                int(r.get("buy_tx", 0) or 0),
                int(r.get("sell_tx", 0) or 0),
                num(r.get("net_shares")),
                num(r.get("net_value_usd")),
                int(r.get("unique_insiders", 0) or 0),
                r.get("first_tx_date", "") or "",
                r.get("last_tx_date", "") or "",
                r.get("error", "") or "",
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        if h in {"entityName", "error"}:
            ws.column_dimensions[col].width = 40
        elif h in {"cik"}:
            ws.column_dimensions[col].width = 12
        else:
            ws.column_dimensions[col].width = 18

    # number formats
    for row in range(2, ws.max_row + 1):
        for col_name in ("net_shares", "net_value_usd"):
            cidx = headers.index(col_name) + 1
            cell = ws.cell(row=row, column=cidx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    wb.save(p)
    log.info("Wrote %s (%d rows)", str(p), len(results))