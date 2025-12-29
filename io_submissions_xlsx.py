from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_submissions_snapshot_xlsx(
    results: list[dict[str, Any]],
    *,
    days: tuple[int, ...],
    path: str = "sec_submissions_features.xlsx",
) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "submissions"

    # dynamic headers by windows
    headers = ["cik", "entityName", "event_date"]
    for nd in days:
        headers += [
            f"eightk_count_{nd}d",
            f"eightk_per_30d_{nd}d",
            f"nt_10k_count_{nd}d",
            f"nt_10q_count_{nd}d",
            f"late_filer_flag_{nd}d",
        ]
    headers += ["days_since_last_10k_or_10q", "error"]

    ws.append(headers)

    # header style
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        row = [r.get(h, "") for h in headers]
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # widths
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        if h in {"entityName", "error"}:
            ws.column_dimensions[col].width = 44
        elif h in {"cik"}:
            ws.column_dimensions[col].width = 12
        else:
            ws.column_dimensions[col].width = 20

    # number formats
    for row_i in range(2, ws.max_row + 1):
        for nd in days:
            per_col = headers.index(f"eightk_per_30d_{nd}d") + 1
            cell = ws.cell(row=row_i, column=per_col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000"

    wb.save(p)


def write_used_dates_xlsx(
    rows: list[dict[str, str]],
    path: str
) -> None:
    """
    Writes the 'used filings' audit log to a separate Excel file.
    """
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "used_dates"

    headers = ["cik", "entityName", "event_date", "filing_date", "form"]
    ws.append(headers)

    # header style
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E0E0E0")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        row = [r.get(h, "") for h in headers]
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # simple widths
    ws.column_dimensions["A"].width = 12  # cik
    ws.column_dimensions["B"].width = 40  # entity
    ws.column_dimensions["C"].width = 15  # event_date
    ws.column_dimensions["D"].width = 15  # filing_date
    ws.column_dimensions["E"].width = 15  # form

    wb.save(p)