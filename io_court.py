from __future__ import annotations

import logging
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


log = logging.getLogger("io_court")


def _to_iso_date(val: Any) -> str | None:
    if val is None:
        return None

    if isinstance(val, (datetime, date)):
        return val.date().isoformat() if isinstance(val, datetime) else val.isoformat()

    s = str(val).strip()
    if not s:
        return None

    # Handle Excel serial dates if they come in as strings/floats (rare but possible)
    if s.replace(".", "", 1).isdigit():
        try:
            # Excel base date usually 1899-12-30
            dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(float(s)) - 2)
            return dt.date().isoformat()
        except Exception:
            pass

    # already ISO
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]

    # Standard US formats found in datasets (MM/DD/YYYY, etc.)
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%m-%d-%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass

    return None


def map_lopucki_to_courtlistener(dist_filed: str) -> str:
    """
    Maps LoPucki 'DistFiled' (e.g. 'NY SD') to CourtListener 'court' (e.g. 'nysb').
    """
    if not dist_filed:
        return ""
    
    clean = str(dist_filed).strip().upper()
    
    # 1. Specific overrides for major bankruptcy venues
    MAPPING = {
        "DE": "deb",        # Delaware
        "NY SD": "nysb",    # NY Southern (Manhattan)
        "NY ED": "nyeb",    # NY Eastern (Brooklyn/Central Islip)
        "TX SD": "txsb",    # TX Southern (Houston)
        "TX ND": "txnb",    # TX Northern (Dallas)
        "TX WD": "txwb",    # TX Western (San Antonio)
        "TX ED": "txeb",    # TX Eastern (Plano)
        "IL ND": "ilnb",    # IL Northern (Chicago)
        "CA CD": "cacb",    # CA Central (LA/Santa Ana)
        "CA ND": "canb",    # CA Northern (San Jose/SF)
        "CA SD": "casb",    # CA Southern (San Diego)
        "VA ED": "vaeb",    # VA Eastern (Richmond)
        "NJ":    "njb",     # New Jersey (District is statewide)
        "NV":    "nvb",     # Nevada
        "MA":    "mab",     # Massachusetts
        "FL SD": "flsb",    # FL Southern (Miami)
        "FL MD": "flmb",    # FL Middle (Tampa/Orlando)
        "MO ED": "moeb",    # MO Eastern (St. Louis)
        "OH SD": "ohsb",    # OH Southern
        "OH ND": "ohnb",    # OH Northern
    }
    
    if clean in MAPPING:
        return MAPPING[clean]

    # 2. Algorithmic Fallback: "STATE DIST" -> "statedistb"
    #    e.g., "PA ED" -> "paeb"
    parts = clean.split()
    
    if len(parts) == 2:
        state, dist = parts
        state = state.lower()
        # LoPucki uses SD, ND, ED, WD, CD, MD
        direction_map = {
            "SD": "s", "ND": "n", "ED": "e", 
            "WD": "w", "CD": "c", "MD": "m"
        }
        if dist in direction_map:
            return f"{state}{direction_map[dist]}b"
            
    # Case: "AZ" (Single district states) -> "azb"
    if len(parts) == 1 and len(clean) == 2:
        return f"{clean.lower()}b"

    # Fallback: remove spaces, lowercase, append 'b' if not present (heuristic)
    # This handles edge cases like "Guam" or similar if they appear.
    val = clean.lower().replace(" ", "")
    if not val.endswith("b"):
        val += "b"
    return val


def clean_docket_number(val: Any) -> str:
    """
    Cleans LoPucki 'CaseNum' to standard docket format.
    LoPucki often has '01-12345' or '01-12345-AJG'.
    CourtListener prefers '01-12345'.
    """
    if val is None:
        return ""
    s = str(val).strip()
    
    # Remove judge initials or extra letters at the end
    # Regex: Look for Year-Number (e.g., 99-12345 or 01-12345) and ignore trailing letters
    # This regex captures "digits-digits" and ignores the rest
    match = re.match(r"^(\d{1,4}-\d+)", s)
    if match:
        return match.group(1)
        
    return s


def load_court_cases_xlsx(path: str, sheet_name: str | None = None) -> list[dict[str, str]]:
    """
    Reads LoPucki BRD Excel format.
    Required Columns (based on your variables list):
      - CikBefore -> cik
      - DistFiled -> court
      - CaseNum   -> docket_number
      - DateFiled -> filed_date
      - NameCorp  -> case_name (optional, for verification)
    """
    if not os.path.exists(path):
        log.error("Excel file not found: %s (cwd=%s)", path, os.getcwd())
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Find headers
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []

    # Normalize headers to lowercase for matching
    header = [str(x).strip().lower() if x is not None else "" for x in header_row]

    def find_col(cands: set[str]) -> int | None:
        for i, h in enumerate(header):
            if h in cands:
                return i
        return None

    # --- Column Mapping based on LoPucki Variables List ---
    cik_col = find_col({"cikbefore", "cik", "company_cik"})
    court_col = find_col({"distfiled", "dist_filed", "court", "district"})
    docket_col = find_col({"casenum", "case_num", "docket", "docket_number"})
    filed_col = find_col({"datefiled", "date_filed", "filed", "petition_date"})
    name_col = find_col({"namecorp", "name_corp", "company_name", "entity_name"})

    missing = []
    if court_col is None: missing.append("DistFiled")
    if docket_col is None: missing.append("CaseNum")
    if filed_col is None: missing.append("DateFiled")

    if missing:
        log.error("Missing columns in %s: %s. (Looking for LoPucki names: DistFiled, CaseNum, DateFiled)", path, missing)
        return []

    out: list[dict[str, str]] = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
            
        # Extract CIK
        cik = ""
        if cik_col is not None and row[cik_col] is not None:
            # Clean CIK (remove decimals if Excel treated as number)
            c_raw = str(row[cik_col]).strip()
            if c_raw.endswith(".0"): 
                c_raw = c_raw[:-2]
            cik = c_raw.zfill(10)

        # Extract Court (DistFiled)
        raw_court = str(row[court_col]).strip() if row[court_col] is not None else ""
        court_code = map_lopucki_to_courtlistener(raw_court)

        # Extract Docket (CaseNum)
        raw_docket = str(row[docket_col]).strip() if row[docket_col] is not None else ""
        docket_number = clean_docket_number(raw_docket)

        # Extract Date (DateFiled)
        filed_date = _to_iso_date(row[filed_col])
        
        # Extract Name (Optional)
        case_name = ""
        if name_col is not None and row[name_col] is not None:
            case_name = str(row[name_col]).strip()

        # Validation: Must have at least docket + date to attempt a search
        if not docket_number or not filed_date:
            continue

        out.append({
            "cik": cik,
            "court": court_code,
            "docket_number": docket_number,
            "filed_date": filed_date,
            "case_name_input": case_name 
        })

    log.info("Loaded %d rows from %s", len(out), path)
    return out


def write_court_metrics_xlsx(results: list[dict[str, Any]], path: str = "court_metrics.xlsx") -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    day_keys: list[str] = []
    if results:
        keys = set().union(*(r.keys() for r in results if isinstance(r, dict)))
        day_keys = sorted([k for k in keys if k.startswith("docket_count_") or k.startswith("motion_count_")])

    headers = [
        "cik",
        "court",
        "docket_number",
        "filed_date",
        "found",
        "docket_id",
        "case_name",        # Found by CourtListener
        "case_name_input",  # From Input File (NameCorp)
        "document_count",
        *day_keys,
        "error",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "court_metrics"
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        if not isinstance(r, dict):
            continue
        ws.append([r.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        if h in {"case_name", "case_name_input", "error"}:
            ws.column_dimensions[col].width = 40
        elif h in {"docket_number"}:
            ws.column_dimensions[col].width = 20
        else:
            ws.column_dimensions[col].width = 15

    wb.save(p)
    log.info("Wrote %s (%d rows)", str(p), len(results))